import json
import xml.etree.ElementTree as ET
import os
import difflib
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font
import google.generativeai as genai
import requests
import re
from collections import Counter
import tempfile

API_KEY = "AIzaSyCOKPBzAeqOocXuiiH44NGBIGLX-IrnlrY"
genai.configure(api_key="AIzaSyAz4cQH5kOVRLqYcMHpU50wgm4Acaqcgj8")
model = genai.GenerativeModel("gemini-pro")

CONFIDENCE_COLORS = {
    "high": "92D050",
    "medium": "FFFF00",
    "low": "FF5050"
}

def extract_json_paths(data, parent_key="", seen_paths=None):
    if seen_paths is None:
        seen_paths = set()
    paths = []
    
    if isinstance(data, dict):
        for key, value in data.items():
            new_key = f"{parent_key}.{key}" if parent_key else key
            if isinstance(value, (dict, list)):
                paths.extend(extract_json_paths(value, new_key, seen_paths))
            else:
                if new_key not in seen_paths:
                    seen_paths.add(new_key)
                    paths.append(new_key)
    elif isinstance(data, list):
        new_key = f"{parent_key}[*]" if parent_key else "[*]"
        if new_key not in seen_paths:
            seen_paths.add(new_key)
            if data:
                paths.extend(extract_json_paths(data[0], new_key, seen_paths))
    
    return paths

def extract_xml_paths(element, parent_path="", seen_paths=None):
    if seen_paths is None:
        seen_paths = set()
    paths = []
    
    current_path = f"{parent_path}/{element.tag}" if parent_path else element.tag
    if list(element):
        for child in element:
            paths.extend(extract_xml_paths(child, current_path, seen_paths))
    else:
        if current_path not in seen_paths:
            seen_paths.add(current_path)
            paths.append(current_path)
    
    return paths

def extract_edifact_fields(content, seg_sep="'", elem_sep="+", sub_elem_sep=":"):
    fields = []
    segment_counts = Counter()
    seen_paths = set()
    
    segments = content.strip().split(seg_sep)
    
    for segment in segments:
        parts = segment.strip().split(elem_sep)  
        segment_type = parts[0].strip()
        segment_counts[segment_type] += 1  

    for segment in segments:
        segment = segment.strip()
        if not segment:
            continue
        
        parts = segment.strip().split(elem_sep)
        segment_type = parts[0].strip()
        qualifier = parts[1] if len(parts) > 1 else ""
       
        for i, element in enumerate(parts[1:], start=1):
            sub_elements = element.split(sub_elem_sep)
            
            field_index = str(i)
            if i < 10:
                field_index = "0" + str(i)

            if len(sub_elements) > 1:
                for j, sub_element in enumerate(sub_elements, start=1):
                    field_path = f"{segment_type}{field_index}.{j}_{qualifier}" if segment_counts[segment_type] > 1 else f"{segment_type}{field_index}.{j}"
                    if field_path not in seen_paths and sub_element.strip():
                        seen_paths.add(field_path)
                        fields.append(field_path)
            else:
                field_path = f"{segment_type}{field_index}_{qualifier}" if segment_counts[segment_type] > 1 else f"{segment_type}{field_index}"
                if field_path not in seen_paths:
                    seen_paths.add(field_path)
                    fields.append(field_path)
    
    return fields

def extract_x12_fields(content, seg_sep="~", elem_sep="*", sub_elem_sep=":"):
    fields = []
    seen_paths = set()
    segment_counts = Counter()

    segments = content.strip().split(seg_sep)
    
    for segment in segments:
        parts = segment.strip().split(elem_sep)
        segment_type = parts[0].strip()
        segment_counts[segment_type] += 1

    for segment in segments:
        segment = segment.strip()
        if not segment:
            continue

        parts = segment.split(elem_sep)  
        segment_type = parts[0].strip() 
        qualifier = parts[1].strip() if len(parts) > 1 else ""

        for i, element in enumerate(parts[1:], start=1):
            sub_elements = element.split(sub_elem_sep) 
            field_index = str(i)
            if i < 10:
                field_index = "0" + str(i)
                
            if len(sub_elements) > 1:
                for j, sub_element in enumerate(sub_elements, start=1):
                    field_path = f"{segment_type}{field_index}.{j}_{qualifier}" if segment_counts[segment_type] > 1 else f"{segment_type}{field_index}.{j}"
                    if field_path not in seen_paths and sub_element.strip():
                        seen_paths.add(field_path)
                        fields.append(field_path)
            else:
                field_path = f"{segment_type}{field_index}_{qualifier}" if segment_counts[segment_type] > 1 else f"{segment_type}{field_index}"
                if field_path not in seen_paths:
                    seen_paths.add(field_path)
                    fields.append(field_path)

    return fields

def read_content(content_or_path, file_type, is_file_path=True, seg_sep="~", elem_sep="*", sub_elem_sep=":"):
    content = ""
    
    # If it's a file path, read the file
    if is_file_path:
        if not os.path.exists(content_or_path):
            raise FileNotFoundError(f"File not found: {content_or_path}")
        
        with open(content_or_path, "r", encoding="utf-8") as file:
            content = file.read()
    else:
        # Use the raw content directly
        content = content_or_path
    
    # Process based on file type
    if file_type == "json":
        try:
            json_data = json.loads(content)
            return extract_json_paths(json_data)
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON format: {e}")
    
    elif file_type == "xml":
        try:
            return extract_xml_paths(ET.fromstring(content))
        except ET.ParseError as e:
            raise ValueError(f"Invalid XML format: {e}")
    
    elif file_type == "edifact":
        return extract_edifact_fields(content, seg_sep, elem_sep, sub_elem_sep)
    
    elif file_type == "x12":
        return extract_x12_fields(content, seg_sep, elem_sep, sub_elem_sep)
    
    else:
        raise ValueError("Unsupported file type. Use JSON, XML, EDIFACT, or X12.")

def call_gemini_api(prompt):
    API_URL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent"

    headers = {"Content-Type": "application/json"}
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.3, "maxOutputTokens": 65536}
    }

    try:
        response = requests.post(f"{API_URL}?key={API_KEY}", headers=headers, json=payload)
        response.raise_for_status()
        response_json = response.json()

        ai_response = response_json["candidates"][0]["content"]["parts"][0]["text"]
        with open("backTick.txt", "r", encoding="utf-8") as file:
            backTickText = file.read()
        cleaned_json = re.sub(fr"{backTickText}json\s*|\s*{backTickText}", "", ai_response).strip()
        return json.loads(cleaned_json)

    except requests.exceptions.RequestException as e:
        print(f"❌ API request failed: {e}")
    except json.JSONDecodeError:
        print(f"❌ Failed to parse JSON from API response")
    except KeyError as e:
        print(f"❌ Missing expected key in API response: {e}")
    except ValueError as e:
        print(f"❌ Validation Error: {e}")

    return {}

def generate_edi_description_prompt(edi_fields):
    with open("ediPrompt.txt", "r", encoding="utf-8") as file:
        prompt = file.read()
    print(f"generate_edi_description_prompt..."+prompt)
    return prompt

def get_edi_field_descriptions(edi_fields):
    description_prompt = generate_edi_description_prompt(edi_fields)
    # print(f"🤖 Generating EDI field descriptions with AI..."+description_prompt)
    return call_gemini_api(description_prompt)

def generate_mapping_prompt(source_fields, target_fields):
    with open("mainPromt.txt", "r", encoding="utf-8") as file:
        prompt_template = file.read()

    # Convert source and target fields to formatted JSON strings
    sourceNames = json.dumps(source_fields, indent=2)
    targetNames = json.dumps(target_fields, indent=2)

    # Replace placeholders with actual values
    prompt = prompt_template.format(sourceNames=sourceNames, targetNames=targetNames)

    print(f"Generated Prompt:\n{prompt}")
    return prompt

def get_mapping_from_ai(source_fields, target_fields):
    try:
        prompt = generate_mapping_prompt(source_fields, target_fields)
        # print(f"🤖 Generating field mappings with AI..."+prompt)
        field_mappings = call_gemini_api(prompt)
    
        if not field_mappings:
            raise ValueError("❌ Error: Gemini API returned an empty mapping!")

        if not isinstance(field_mappings, dict):
            raise TypeError("❌ Error: Gemini API response is not a dictionary!")

        return field_mappings

    except Exception as e:
        raise RuntimeError(f"🔥 Error in get_mapping_from_ai: {e}")

def normalize_confidence(confidence):
    if confidence is None:
        return 0.0  
    elif confidence > 1:
        return round(confidence, 2)
    else:
        return round(confidence * 100, 2)

def find_best_match(target_field, source_fields):
    best_match = None
    best_score = 0
    for source_field in source_fields:
        score = difflib.SequenceMatcher(None, target_field.lower(), source_field.lower()).ratio()
        if score > best_score:
            best_score = score
            best_match = source_field
    return best_match, round(best_score * 100, 2)

def create_excel_mapping(source_fields, target_fields, field_mappings, excel_filename, source_format, target_format, seg_sep, elem_sep, sub_elem_sep):
    wb = Workbook()
    ws = wb.active
    ws.title = "Field Mapping"
    headers = ["Target Field", "Source Field (Dropdown)", "Confidence (%)", "Mapping Type", "Logic"]
    ws.append(headers)
    
    bold_font = Font(bold=True)
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header).font = bold_font

    mapping_type_dv = DataValidation(type="list", formula1='"Direct,Logic"', showDropDown=True)

    ws_hidden = wb.create_sheet(title="DropdownValues")
    extra_options = ["<Not Mapped>", "<Constant>", "<No Mapping Needed>", "<Logic>"]

    high_conf_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    medium_conf_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    low_conf_fill = PatternFill(start_color="FF5050", end_color="FF5050", fill_type="solid")

    ws_hidden["A1"] = "Source Fields"
    actual_source_fields = [str(field) for field in source_fields]
    actual_target_fields = [str(field) for field in target_fields]
    total_dropdown_fields = extra_options + actual_source_fields

    for i, path in enumerate(total_dropdown_fields, start=2):
        ws_hidden[f"A{i}"] = path

    dropdown_range = f"'DropdownValues'!$A$2:$A${len(total_dropdown_fields) + 1}"

    if not isinstance(field_mappings, dict):
        raise TypeError("❌ Error: field_mappings should be a dictionary!")

    for row_num, target_field in enumerate(target_fields, start=2):
        mapping = field_mappings.get(target_field, None)
        if not isinstance(mapping, dict):
            mapping = {}
        
        if mapping:
            source_field = mapping.get("source_field", "<Not Mapped>")
            confidence = normalize_confidence(mapping.get("confidence", 0))
            mapping_type = mapping.get("mapping_type", "Direct")
            logic = mapping.get("logic", "")
        else:
            source_field = "<Not Mapped>"
            confidence = 0
            mapping_type = "Manual"
            logic = "Needs manual review"

        dv = DataValidation(type="list", formula1=dropdown_range, showDropDown=True)
        ws.add_data_validation(dv)

        source_cell = ws.cell(row=row_num, column=2)
        if isinstance(source_field, list):
            source_cell.value = ", ".join(source_field)
        else:
            source_cell.value = source_field if source_field else "Not Mapped"
        dv.add(source_cell)

        if confidence >= 80:
            source_cell.fill = high_conf_fill
        elif confidence >= 50:
            source_cell.fill = medium_conf_fill
        else:
            source_cell.fill = low_conf_fill

        ws.cell(row=row_num, column=1, value=target_field)
        ws.cell(row=row_num, column=3, value=confidence)
        
        mapping_type_cell = ws.cell(row=row_num, column=4, value=mapping_type)
        mapping_type_dv.add(mapping_type_cell)
        ws.add_data_validation(mapping_type_dv)

        ws.cell(row=row_num, column=5, value=logic)

    ws_hidden.sheet_state = "hidden"

    if source_format in ["edifact", "x12"] or target_format in ["edifact", "x12"]:
        if source_format in ["edifact", "x12"]:
            edi_fields = source_fields
        else:
            edi_fields = target_fields
            
        ws_desc = wb.create_sheet(title="EDI Field Descriptions")
        ws_desc.append(["Field Name", "Description"])

        edi_descriptions = get_edi_field_descriptions(edi_fields)

        if edi_descriptions:
            for field, description in edi_descriptions.items():
                ws_desc.append([field, description])

    wb.save(excel_filename)
    print(f"✅ Excel file '{excel_filename}' created successfully! Open it to review mappings.")

def detect_format_from_content(content):
    """Attempt to auto-detect the format of the content"""
    content = content.strip()
    
    # Check for JSON
    if (content.startswith('{') and content.endswith('}')) or (content.startswith('[') and content.endswith(']')):
        try:
            json.loads(content)
            return "json"
        except json.JSONDecodeError:
            pass
    
    # Check for XML
    if content.startswith('<?xml') or (content.startswith('<') and content.endswith('>')):
        try:
            ET.fromstring(content)
            return "xml"
        except ET.ParseError:
            pass
    
    # Check for EDIFACT (starts with UNB or UNA)
    if 'UNB' in content[:20] or 'UNA' in content[:20]:
        return "edifact"
    
    # Check for X12 (often starts with ISA)
    if content.startswith('ISA'):
        return "x12"
    
    # Default to requiring user input
    return None

def save_raw_content_to_temp_file(content, extension=".txt"):
    """Save raw content to a temporary file and return the file path"""
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=extension)
    with open(temp_file.name, 'w', encoding='utf-8') as f:
        f.write(content)
    return temp_file.name

def get_user_input(auto_detect=True):
    valid_formats = ["json", "xml", "edifact", "x12"]
    # source_is_file = input("Is your source a file path? (y/n): ").strip().lower() == 'y'
    source_is_file = False
    
    if source_is_file:
        source_file_path = input("Enter source file path: ").strip()
        source_format = input("Enter source format (JSON/XML/EDIFACT/X12): ").strip().lower()
        source_content = None
    else:
        print("Enter/paste source content (press Ctrl+D on Unix/Linux or Ctrl+Z followed by Enter on Windows to finish):")
        source_content = ""
        while True:
            try:
                line = input()
                source_content += line + "\n"
            except EOFError:
                break
        
        if auto_detect:
            detected_format = detect_format_from_content(source_content)
            if detected_format:
                print(f"Auto-detected source format: {detected_format.upper()}")
                source_format = detected_format
            else:
                source_format = input("Enter source format (JSON/XML/EDIFACT/X12): ").strip().lower()
        else:
            source_format = input("Enter source format (JSON/XML/EDIFACT/X12): ").strip().lower()
        
        source_file_path = None
    
    while source_format not in valid_formats:
        print("Invalid format. Please enter JSON, XML, EDIFACT, or X12.")
        source_format = input("Enter source format (JSON/XML/EDIFACT/X12): ").strip().lower()

    # target_is_file = input("Is your target a file path? (y/n): ").strip().lower() == 'y'
    target_is_file = False
    
    if target_is_file:
        target_file_path = input("Enter target file path: ").strip()
        target_format = input("Enter target format (JSON/XML/EDIFACT/X12): ").strip().lower()
        target_content = None
    else:
        print("Enter/paste target content (press Ctrl+D on Unix/Linux or Ctrl+Z followed by Enter on Windows to finish):")
        target_content = ""
        while True:
            try:
                line = input()
                target_content += line + "\n"
            except EOFError:
                break
        
        if auto_detect:
            detected_format = detect_format_from_content(target_content)
            if detected_format:
                print(f"Auto-detected target format: {detected_format.upper()}")
                target_format = detected_format
            else:
                target_format = input("Enter target format (JSON/XML/EDIFACT/X12): ").strip().lower()
        else:
            target_format = input("Enter target format (JSON/XML/EDIFACT/X12): ").strip().lower()
        
        target_file_path = None
    
    while target_format not in valid_formats:
        print("Invalid format. Please enter JSON, XML, EDIFACT, or X12.")
        target_format = input("Enter target format (JSON/XML/EDIFACT/X12): ").strip().lower()

    if source_format in ["edifact", "x12"] or target_format in ["edifact", "x12"]:
        default_seg_sep = "'" if source_format == "edifact" or target_format == "edifact" else "~"
        default_elem_sep = "+" if source_format == "edifact" or target_format == "edifact" else "*"
        seg_sep = input(f"Enter segment separator (default: {default_seg_sep}): ").strip() or default_seg_sep
        elem_sep = input(f"Enter element separator (default: {default_elem_sep}): ").strip() or default_elem_sep
        sub_elem_sep = input("Enter sub-element separator (default: :): ").strip() or ":"
    else:
        seg_sep, elem_sep, sub_elem_sep = "'", "+", ":"
    
    return {
        "source_format": source_format,
        "source_file_path": source_file_path,
        "source_content": source_content,
        "source_is_file": source_is_file,
        "target_format": target_format,
        "target_file_path": target_file_path,
        "target_content": target_content,
        "target_is_file": target_is_file,
        "seg_sep": seg_sep,
        "elem_sep": elem_sep,
        "sub_elem_sep": sub_elem_sep
    }

def create_field_mapping(source_file_type=None, source_file=None, target_file_type=None, target_file=None, auto_detect=False, seg_sep="'", elem_sep="+", sub_elem_sep=":"):
    """
    Create field mappings between source and target files with optional format auto-detection.
    
    Parameters:
    source_file_type (str): Format of source file (json, xml, edifact, x12) or None for auto-detection
    source_file (str): Content or file path of source file
    target_file_type (str): Format of target file (json, xml, edifact, x12) or None for auto-detection
    target_file (str): Content or file path of target file
    auto_detect (bool): Whether to auto-detect file formats if not specified
    seg_sep (str): Segment separator for EDI formats
    elem_sep (str): Element separator for EDI formats
    sub_elem_sep (str): Sub-element separator for EDI formats
    
    Returns:
    str: Path to the created Excel file with field mappings
    """
    valid_formats = ["json", "xml", "edifact", "x12"]
    source_is_file = os.path.exists(source_file) if source_file else False
    target_is_file = os.path.exists(target_file) if target_file else False
    
    # Handle source format auto-detection if needed
    if auto_detect and source_file_type is None and source_file:
        if source_is_file:
            with open(source_file, 'r', encoding='utf-8') as f:
                content = f.read()
            source_file_type = detect_format_from_content(content)
        else:
            source_file_type = detect_format_from_content(source_file)
            
    # Ensure source format is valid
    if source_file_type not in valid_formats:
        raise ValueError(f"Invalid source format. Please use one of: {', '.join(valid_formats)}")
    
    # Handle target format auto-detection if needed
    if auto_detect and target_file_type is None and target_file:
        if target_is_file:
            with open(target_file, 'r', encoding='utf-8') as f:
                content = f.read()
            target_file_type = detect_format_from_content(content)
        else:
            target_file_type = detect_format_from_content(target_file)
    
    # Ensure target format is valid
    if target_file_type not in valid_formats:
        raise ValueError(f"Invalid target format. Please use one of: {', '.join(valid_formats)}")
    
    # Set appropriate separators for EDI formats if needed
    if source_file_type in ["edifact", "x12"] or target_file_type in ["edifact", "x12"]:
        if source_file_type == "edifact" or target_file_type == "edifact":
            seg_sep = seg_sep or "'"
            elem_sep = elem_sep or "+"
        elif source_file_type == "x12" or target_file_type == "x12":
            seg_sep = seg_sep or "~"
            elem_sep = elem_sep or "*"
        sub_elem_sep = sub_elem_sep or ":"
    
    # Process source fields
    print(f"Processing source in {source_file_type.upper()} format...")
    source_fields = read_content(
        source_file, 
        source_file_type, 
        is_file_path=source_is_file,
        seg_sep=seg_sep, 
        elem_sep=elem_sep, 
        sub_elem_sep=sub_elem_sep
    )
    
    # Process target fields
    print(f"Processing target in {target_file_type.upper()} format...")
    target_fields = read_content(
        target_file, 
        target_file_type, 
        is_file_path=target_is_file,
        seg_sep=seg_sep, 
        elem_sep=elem_sep, 
        sub_elem_sep=sub_elem_sep
    )
    
    print(f"Found {len(source_fields)} source fields")
    print(f"Found {len(target_fields)} target fields")
    
    # Generate AI mappings
    print("Generating field mappings with AI...")
    field_mappings = get_mapping_from_ai(source_fields, target_fields)
    
    if not field_mappings:
        print("⚠️ Warning: field_mappings is empty. No mappings to process!")
    else:
        print(f"Generated {len(field_mappings)} field mappings")
    
    # Create Excel file with mappings
    excel_filename = "AI_Field_Mapping.xlsx"
    create_excel_mapping(
        source_fields, 
        target_fields, 
        field_mappings, 
        excel_filename, 
        source_file_type, 
        target_file_type, 
        seg_sep, 
        elem_sep, 
        sub_elem_sep
    )
    
    return excel_filename

# Update main function to use the new create_field_mapping function
def main(source_file_type=None, source_file=None, target_file_type=None, target_file=None, auto_detect=False):
    # Get user input with optional format auto-detection
    # use_function = input("Use direct function call? (y/n): ").strip().lower() == 'y'
    use_function = True
    
    if use_function:
        # source_file_type = input("Enter source format (JSON/XML/EDIFACT/X12) or leave empty for auto-detect: ").strip().lower() or None
        # source_file = input("Enter source file path or paste content: ").strip()
        
        # target_file_type = input("Enter target format (JSON/XML/EDIFACT/X12) or leave empty for auto-detect: ").strip().lower() or None
        # target_file = input("Enter target file path or paste content: ").strip()
        
        auto_detect = source_file_type is None or target_file_type is None
        
        if "edifact" in [source_file_type, target_file_type] or "x12" in [source_file_type, target_file_type]:
            seg_sep = input("Enter segment separator (default depends on format): ").strip() or None
            elem_sep = input("Enter element separator (default depends on format): ").strip() or None
            sub_elem_sep = input("Enter sub-element separator (default: :): ").strip() or ":"
        else:
            seg_sep, elem_sep, sub_elem_sep = "'", "+", ":"
        
        excel_file = create_field_mapping(
            source_file_type=source_file_type,
            source_file=source_file,
            target_file_type=target_file_type,
            target_file=target_file,
            auto_detect=auto_detect,
            seg_sep=seg_sep,
            elem_sep=elem_sep,
            sub_elem_sep=sub_elem_sep
        )
        
        print(f"✅ Excel file '{excel_file}' created successfully! Open it to review mappings.")
    else:
        # Use the original interactive input method
        user_input = get_user_input(auto_detect=True)
        
        print(f"Processing with separators: {user_input['seg_sep']} | {user_input['elem_sep']} | {user_input['sub_elem_sep']}")
        
        # Process source
        if user_input["source_is_file"]:
            source_fields = read_content(
                user_input["source_file_path"], 
                user_input["source_format"], 
                is_file_path=True,
                seg_sep=user_input["seg_sep"], 
                elem_sep=user_input["elem_sep"], 
                sub_elem_sep=user_input["sub_elem_sep"]
            )
        else:
            source_fields = read_content(
                user_input["source_content"], 
                user_input["source_format"], 
                is_file_path=False,
                seg_sep=user_input["seg_sep"], 
                elem_sep=user_input["elem_sep"], 
                sub_elem_sep=user_input["sub_elem_sep"]
            )
        
        # Process target
        if user_input["target_is_file"]:
            target_fields = read_content(
                user_input["target_file_path"], 
                user_input["target_format"], 
                is_file_path=True,
                seg_sep=user_input["seg_sep"], 
                elem_sep=user_input["elem_sep"], 
                sub_elem_sep=user_input["sub_elem_sep"]
            )
        else:
            target_fields = read_content(
                user_input["target_content"], 
                user_input["target_format"], 
                is_file_path=False,
                seg_sep=user_input["seg_sep"], 
                elem_sep=user_input["elem_sep"], 
                sub_elem_sep=user_input["sub_elem_sep"]
            )

        print(f"Found {len(source_fields)} source fields")
        print(f"Found {len(target_fields)} target fields")

        field_mappings = get_mapping_from_ai(source_fields, target_fields)
        
        if not field_mappings:
            print("⚠️ Warning: field_mappings is empty. No mappings to process!")
        else:
            print(f"Generated {len(field_mappings)} field mappings")
            
        excel_filename = "AI_Field_Mapping.xlsx"
        create_excel_mapping(
            source_fields, 
            target_fields, 
            field_mappings, 
            excel_filename, 
            user_input["source_format"], 
            user_input["target_format"], 
            user_input["seg_sep"], 
            user_input["elem_sep"], 
            user_input["sub_elem_sep"]
        )

if __name__ == "__main__":
    main(source_file_type="json", source_file="source.json", target_file_type="json", target_file="target.json")